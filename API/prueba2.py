from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.formatting.rule import Rule, IconSet, FormatObject

# Crear un DataFrame de ejemplo

fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
text_color = Font(bold=True, color='FFFFFF')
alignment_center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
alignment_center_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_center_left = Alignment(horizontal="left", vertical="top", wrap_text=True)


type_font = font = Font(
    name='Calibri',  # Tipo de fuente (por defecto es 'Calibri')
    size=12,         # Tamaño de la fuente
    bold=True,      # Negrita (True para habilitarla)
    color='FFFFFF'   # Color de la fuente en formato HEX (e.g., 'FF0000' para rojo)
)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wb = load_workbook("informe_diciembre.xlsx")
ws = wb.active

ws.insert_rows(1, amount = 3)

ws.merge_cells('F3:H3')

ws['F3'].fill = fill_color
ws['F3'].font = text_color
ws['F3'].alignment = alignment_center_top
ws['F3'].value = "Año 2024"


for col in ws.iter_cols(min_col = 0, max_col = 13, min_row = 4 , max_row = 4):
    for cell in col: 
        cell.font = text_color
        cell.fill = fill_color
        cell.font = type_font
        cell.alignment = alignment_center_center
        
for col in ws.iter_cols(min_col = 0, max_col = 13, min_row = 5, max_row = 111):
    for cell in col:
        cell.alignment = alignment_center_left
        cell.border = thin_border
        
for col in ws.iter_cols(min_col = 5, max_col = 12, min_row = 5 , max_row = 111):
    for cell in col:
        cell.alignment = alignment_center_top


list_range_columns = []
for col in ws.iter_cols(min_col = 0, max_col = 13, min_row = 4, max_row = 4):
    for cell in col:
        list_range_columns.append(len(cell.value))
            

for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        for cell in col:
            if col_letter == 'B': 
                ws.column_dimensions[col_letter].width = list_range_columns[1] + 1
            if col_letter == 'A':
                ws.column_dimensions[col_letter].width = list_range_columns[0] + 8
            if col_letter == 'C':
                ws.column_dimensions[col_letter].width = list_range_columns[2] - 8
            if col_letter == 'E':
                ws.column_dimensions[col_letter].width = list_range_columns[4] + 3
            if col_letter == 'D':
                ws.column_dimensions[col_letter].width = list_range_columns[3] + 1.5
            if col_letter == 'F':
                ws.column_dimensions[col_letter].width = list_range_columns[5] + 3
            if col_letter == 'J':
                ws.column_dimensions[col_letter].width = list_range_columns[9] - 4
            if col_letter == 'G':
                ws.column_dimensions[col_letter].width = list_range_columns[6] + 3
            if col_letter == 'H':
                ws.column_dimensions[col_letter].width = list_range_columns[7]  + 3.5
            if col_letter == 'K':
                ws.column_dimensions[col_letter].width = list_range_columns[10] - 4
            if col_letter == 'I':
                ws.column_dimensions[col_letter].width = list_range_columns[8] - 6.5
            if col_letter == 'L':
                ws.column_dimensions[col_letter].width = list_range_columns[11] - 9
            if col_letter == 'M':
                ws.column_dimensions[col_letter].width = 122
                

for row in ws.iter_rows(min_col = 0, max_col = 13, min_row = 5, max_row = 111):
    max_saturation = 1.0
    for cell in row:
        saturation = ws.column_dimensions[cell.column_letter].width/len(str(cell.value))
        if saturation < max_saturation:
            max_saturation = saturation
            col_saturation = cell.column_letter
            row_saturation = cell.row
            print(max_saturation)
            print(f"{col_saturation}- {row_saturation}")
    #ws.row_dimensions[row[0].row].height = ws.row_dimensions[row[col_saturation].row].height


cfvo_red = FormatObject(type='num', val=3, gte=True)    # Rojo: >= 3
cfvo_yellow = FormatObject(type='num', val=2, gte=True) # Amarillo: >= 2
cfvo_green = FormatObject(type='num', val=2, gte=True) # Verde: < 2

# Configurar los íconos de semáforo en orden invertido
icon_set = IconSet(iconSet='3TrafficLights1', cfvo=[cfvo_green, cfvo_yellow, cfvo_red],reverse = True)

# Crear la regla de formato condicional
rule = Rule(type='iconSet', iconSet=icon_set)
   
   # Aplicar la regla al rango deseado
ws.conditional_formatting.add("K5:K111", rule)



ws.merge_cells('J4:K4')
                       
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Landscape orientation
ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL   
ws.page_setup.scale = 69  # Fit to one page in width

# Set print titles (repeat first row on every printed page)
ws.print_title_rows = '3:4'

# Set margins
ws.page_margins = PageMargins(
    left=0, right=0,
    top=0, bottom=0,
    header=0, footer=0
)

# Set gridlines to be printed
ws.sheet_view.showGridLines = True

        
wb.save("archivo_modificado.xlsx")
