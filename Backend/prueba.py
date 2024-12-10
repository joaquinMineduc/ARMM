from openpyxl import load_workbook
from eval_functions import df_eval_NC


file_path = "APP/Backend/Input/formato.xlsx"
workbook = load_workbook(file_path)

sheet_name = "Ev Proveedores"
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook!")

